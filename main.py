import openpyxl, re, os, time, datetime, pyinputplus as pyip

# Ensures that any file passed to this function has a valid .xlsx extension
def getExcelFileFormat(file):
    if ".xlsx" not in file:
        newFile = file + ".xlsx"
        return newFile
    return file

# Reads the raw applicant data, runs it against business rules, and returns a list of valid rows
def readExcel():
    # Gets name of file user wants to clean
    file = pyip.inputStr("Enter the file name you wish to clean \n")
    # Checks if file is valid and in the same directory as this script
    try:
        applicants = openpyxl.load_workbook(getExcelFileFormat(file))
        sheet = applicants.active
    except:
        print("Something went wrong! (Ensure the excel file you wish to clean is in the same directory as this script!)")
        return
    # Array to hold the successfully cleansed data
    cleanData = []
    # Loop through the excel sheet rows
    for i in range(2, sheet.max_row):
        name = sheet["A" + str(i)].value
        dateSumbitted = sheet["B" + str(i)].value
        programAppliedTo = sheet["C" + str(i)].value
        accepted = sheet["D" + str(i)].value
        email = sheet["E" + str(i)].value
        streetAddress = sheet["F" + str(i)].value
        city = sheet["G" + str(i)].value
        province = sheet["H" + str(i)].value
        postalCode = sheet["I" + str(i)].value
        
        # Run the row data against all validation rules.
        # If any check fails log the error and flag the row as invalid.
        isRowValid = True
        if not validateNoNumbersInString(name):
            createErrorLog(False, name, "A", str(i))
            isRowValid = False
        if not validateDate(dateSumbitted):
            createErrorLog(False, dateSumbitted, "B", str(i))
            isRowValid = False
        if not validateProgram(programAppliedTo):
            createErrorLog(False, programAppliedTo, "C", str(i))
            isRowValid = False
        if not validateAccepted(accepted):
            createErrorLog(False, accepted, "D", str(i))
            isRowValid = False
        if not validateEmail(email):
            createErrorLog(False, email, "E", str(i))
            isRowValid = False
        if not validateCanadianPostalCode(postalCode):
            createErrorLog(False, postalCode, "I", str(i))
            isRowValid = False

        # If the row passes all checks, bundle the data and add it to the clean list
        if isRowValid:
            goodData = [name, dateSumbitted, programAppliedTo, accepted, email, streetAddress, city, province, postalCode]
            cleanData.append(goodData)
    # Close the workbook to free up system memory
    applicants.close()
    # Return data to be used in exportCleanedData()
    return cleanData
            
        
# Outputs all invalid data to an error.txt file for human review        
def createErrorLog(isValid, data, column, row):
    file = "error.txt"
    if not isValid:
        errorMessage = "Invalid data found: \n" + str(data) + " is Located at row: " + str(row) + " column:  " + str(column) + "\n"
        # Ensures that a new error.txt is created if it doesn't exist, otherwise appends
        if not os.path.exists(file):
            f = open(file, "w")
            f.write(errorMessage)
            f.close()
        else:
            f = open(file, "a")
            f.write(errorMessage)
            f.close()
        

# Validates that the applicant has officially accepted their offer
def validateAccepted(acceptedInput):
    if acceptedInput == "Yes":
        return True
    return False

# Ensures the application was submitted for the current academic year (2026 or later)
def validateDate(dateInput):
    try:
       if dateInput.year < 2026:
           return False
    except:
        print("Invalid date: " + str(dateInput))
        return False
    return True

# Verifies that the input string (such as a name) contains only alphabetical characters,
# spaces, hyphens, and apostrophes — handles names like "O'Brien" or "Mary-Jane"
def validateNoNumbersInString(textInput):
    invalidCharRegex = re.compile(r"[^a-zA-Z\s\-']")
    if invalidCharRegex.search(str(textInput)):
        return False
    return True

# Enforces local-part @ domain . tld structure
# Allows dots/hyphens/underscores in the local part and domain, requires a 2+ char TLD
def validateEmail(emailInput):
    emailRegex = re.compile(r'^[\w\.\-]+@[\w\.\-]+\.[a-zA-Z]{2,}$')
    if emailRegex.search(str(emailInput)):
        return True
    return False

# Validates that the postal code strictly matches Canadian alphanumeric format:
# letter-digit-letter space digit-letter-digit (e.g. "A1B 2C3")
# Anchored with ^ and $ to prevent partial matches from passing
def validateCanadianPostalCode(postalInput):
    postalRegex = re.compile(r'^[A-Za-z]\d[A-Za-z]\s\d[A-Za-z]\d$')
    if postalRegex.search(str(postalInput)):
        return True
    return False

# Validates that the applicant is applying strictly to the Computer Programming department
# Anchored with ^ to guard against leading whitespace like "  COMP101" slipping through
def validateProgram(programInput):
    programRegex = re.compile(r'^COMP')
    if programRegex.search(str(programInput)):
        return True
    return False


# Orchestrates the main program loop, prompting the user for files, 
# triggering the cleaning process, and handling the final export.
def main():
    userWantsToContinue = "yes"

    # Loop allows the user to process multiple files in one session
    while not userWantsToContinue == "no":
        # Clean up any previous runs
        if os.path.exists("error.txt"):
            os.remove("error.txt")
        fileName = getExcelFileFormat(pyip.inputStr("What would you like the cleaned data file name to be? \n"))
        if os.path.exists(fileName):
            os.remove(fileName)
        validData = readExcel()
        # Only run the export function if validData actually contains a list
        if validData != None:
            exportCleanedData(validData, fileName)
        userWantsToContinue = pyip.inputYesNo("Do you wish to continue cleaning files? (Enter yes or no)\n")


# Creates a new Excel workbook, writes the header row, 
# and appends all validated student data before saving and calculating the export time.
def exportCleanedData(data, fileName):
    startTime = time.time()
    print("Starting data export...")
    outputWb = openpyxl.Workbook()
    sheet = outputWb.active
    headers = ["Name", "Date Sumbitted", "Program", "Accepted", "Email", "Street Address", "City", "Province", "Postal Code"]
    sheet.append(headers)
    for row in data:
        sheet.append(row)
    
    outputWb.save(fileName)
    outputWb.close()
    endTime = time.time()
    totalTimeTook = endTime - startTime
    print("Data export completed! time took: " + str(totalTimeTook) + " Seconds")


# Call main to execute the application
main()
